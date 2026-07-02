#!/usr/bin/env python3
"""
Build presentation/RIOS-module-comparison.xlsx — a leadership benchmark of every
RIOS module against the reinsurance / insurance software the market actually uses.

Reproducible generator. Run:  python3 presentation/build_comparison.py

HONESTY BASIS (see the "Legend & Honesty Note" sheet in the output):
  * Every RIOS claim is grounded in the repository: web/src/app/nav.ts (module IA),
    server/src/modules/* (endpoints/tables), packages/domain/src/* (the maths that
    exists), and the honest status registers docs/phases.md, docs/open-questions.md,
    docs/industry-gap-analysis.md. No RIOS feature is invented.
  * Status is one of Delivered / Partial / Designed-for, cross-checked against those
    registers.
  * Comparator internals are NOT invented. They are inferred from public / market-
    standard footprints only (vendor module maps, Lloyd's/LMA MRC, ACORD EBOT/ECOT,
    NAIC Schedule F, Solvency II QRTs, published product-model descriptions).

Named comparator set (20): SAP FS-RI, Sapiens ReinsuranceMaster, Eurobase Synergy2,
Guidewire, Duck Creek, Fadata INSIS, Oracle Insurance, FIS, Majesco, SICS (Insurdata/
SICS), Xuber/DXC, Verisk/RMS, Moody's RMS, Sequel (Verisk Sequel), Effisoft (WebXL/
Omega), Tia/TietoEVRY, msg global, Novidea, Instanda, ClarionDoor.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # deep navy
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
NOTE_FONT = Font(italic=True, size=10, color="404040")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILL = {
    "Delivered":    PatternFill("solid", fgColor="C6EFCE"),  # green
    "Partial":      PatternFill("solid", fgColor="FFEB9C"),  # amber
    "Designed-for": PatternFill("solid", fgColor="F8CBAD"),  # orange
    "Not-yet":      PatternFill("solid", fgColor="FFC7CE"),  # red
}


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def finalize(ws, widths, wrap_from_row=2, status_col=None, freeze="A2"):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=wrap_from_row):
        for cell in row:
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if status_col and cell.column == status_col:
                fill = STATUS_FILL.get(str(cell.value).strip())
                if fill:
                    cell.fill = fill
                    cell.alignment = CENTER
    ws.freeze_panes = freeze


# ---------------------------------------------------------------------------
# SHEET 1 — Module Comparison
# columns: Module | Category | How it works in RIOS (core) | Status in RIOS |
#          What makes RIOS's version strong | Comparable RI software |
#          Same module exists? | How the comparators do it (core) | Notes / gaps
# ---------------------------------------------------------------------------
MODULE_HEADERS = [
    "Module", "Category", "How it works in RIOS (core)", "Status in RIOS",
    "What makes RIOS's version strong", "Comparable RI software",
    "Same module exists in those?", "How the comparators do it (core)",
    "Notes / gaps",
]

MODULES = [
    # --- Overview / cross-cutting ---
    ("Dashboard", "Platform",
     "Tenant-scoped KPI summary (/api/dashboard/summary) aggregating live figures from the reconciling financial-event chain.",
     "Partial",
     "Every number is RLS-scoped and traces to the same audited financial events as the GL; no separate reporting store to drift.",
     "SAP FS-RI, Sapiens, Guidewire, Duck Creek, Fadata, Oracle, Majesco",
     "Yes",
     "All primary/reinsurance suites ship configurable role-based home dashboards over a BI/reporting layer.",
     "Drag-and-drop dashboard designer is designed-for (open-questions §9)."),
    ("Executive", "Analytics & Compliance",
     "Executive/board view (server executive module) rolling up portfolio, premium, claims and result KPIs.",
     "Partial",
     "Derived from the reconciling chain, so board figures tie back to postings.",
     "SAP FS-RI, Sapiens, Moody's RMS, FIS, Oracle",
     "Yes",
     "Executive dashboards sit on the vendor BI/analytics layer or an external DW.",
     "Full pivot/drill-through BI is designed-for."),
    ("Intelligence", "AI",
     "Deterministic insight surface (server intelligence module) summarising book signals from live data.",
     "Partial",
     "Transparent, rule-based signals rather than a black box.",
     "Verisk/Sequel, Moody's RMS, Guidewire (analytics add-ons)",
     "Varies",
     "Where present, insight/analytics is a bolt-on BI/ML module, not universal in RI admin systems.",
     "LLM narration is optional and designed-for (open-questions §13)."),
    ("AI Insights", "AI",
     "Renewal-likelihood scoring and portfolio insights via a transparent, unit-tested heuristic (server aiInsights; domain prediction/insight).",
     "Partial",
     "No black box: the score is an inspectable, unit-tested heuristic; an LLM may optionally narrate it.",
     "Verisk, Moody's RMS, Guidewire",
     "Varies",
     "Native predictive scoring is uncommon in RI admin platforms; analytics vendors supply ML separately.",
     "Optional LLM enrichment designed-for; core heuristic delivered."),
    ("Search", "Platform",
     "Global + natural-language search (server search/searchEnhanced; domain nlSearch) over tenant entities.",
     "Partial",
     "Deterministic NL query parsing, RLS-scoped results.",
     "SAP FS-RI, Sapiens, Guidewire, Duck Creek",
     "Yes",
     "Enterprise search over an index (often ElasticSearch/OpenSearch) is standard.",
     "ElasticSearch/OpenSearch backing index is designed-for (open-questions §7)."),
    ("Mobile", "Platform",
     "Condensed responsive projection + PWA manifest (server mobile module) served to the web client.",
     "Partial",
     "Reuses the same RLS-scoped APIs; no separate mobile data path.",
     "Guidewire, Duck Creek, Majesco, Instanda",
     "Partial",
     "Primary-insurance suites offer mobile apps; dedicated RI admin systems rarely do.",
     "Native mobile shell is explicitly out of scope (open-questions §13)."),

    # --- Underwriting ---
    ("Underwriting Workspace", "Underwriting",
     "Submission lifecycle stage machine (Submission->Triage->...->Bound/Declined) with a factor-by-factor risk score; illegal transitions rejected (server underwriting; domain underwriting).",
     "Delivered",
     "Metadata-driven model catalogue (8 structures x 13 LOBs), transparent risk score, illegal-transition guard, full audited trail.",
     "Sequel, Eurobase Synergy2, SAP FS-RI, Sapiens",
     "Partial",
     "Reinsurance UW workbenches with referral/authority exist in specialist suites; breadth varies by product.",
     "Live commercial cat-model API and real OCR extraction connect behind existing adapters (phases.md)."),
    ("Treaty Workspace", "Underwriting",
     "Treaty place->bind lifecycle (server treaties); bind books the deposit premium as a financial event; commercial terms are a typed, validated schema (30+ keys).",
     "Delivered",
     "Binding starts the reconciling event->statement->GL chain; integer money; typed terms validation; lifecycle guard.",
     "SAP FS-RI, Sapiens ReinsuranceMaster, Eurobase Synergy2, Effisoft Omega, DXC",
     "Yes",
     "Treaty administration (register, terms, technical accounting) is the core of every RI admin platform.",
     "Multi-section treaties and leader/follower panel terms remain follow-ons (gap-analysis §1.1)."),
    ("Facultative Workspace", "Underwriting",
     "Facultative contract capture + admin sub-forms (server facultative/facultativeAdmin) on the shared party/contract schema.",
     "Partial",
     "Same typed-terms + reconciling chain as treaty; fac modelled in schema (contract_kind/direction).",
     "Sequel, Eurobase, SAP FS-RI, Sapiens",
     "Yes",
     "Fac administration incl. single-risk fast cession is standard in specialist RI suites.",
     "Single-risk fast cession and fac-obligatory are designed-for (open-questions §1); some API fields lack UI (gap-analysis §1.4)."),
    ("Placement", "Underwriting",
     "Slip signing (POST /api/placement/slips/:id/sign) with sign-down guards (signed<=written, Sum signed<=order) or PRO_RATA auto sign-down, plus written-vs-signed reconciliation.",
     "Delivered",
     "Integer-exact sign-down with hard guards and a written-vs-signed tie-out; participations in schema.",
     "Sequel, Whitespace/PPL-style placing, Eurobase, DXC",
     "Partial",
     "Electronic placing (PPL/Whitespace) and MRC slip management are the London-market standard; fewer non-London vendors cover it.",
     "Leader/follower terms on the panel are a follow-on (gap-analysis Tier 1)."),
    ("Pricing", "Underwriting",
     "Technical/burn-cost premium, combined-ratio what-if (rate-change x loss-shock grid, sensitivity) and swing rating (server pricing; domain pricing/rating/pricingScenarios).",
     "Delivered",
     "Pure, unit-tested domain math; scenario grid and sensitivity in the drawer; swing rating collared to [min,max].",
     "Sequel, Moody's RMS, Verisk, actuarial tooling (Igloo-class)",
     "Partial",
     "Exposure/experience rating lives in specialist actuarial or cat tools; RI admin systems price more lightly.",
     "Stochastic/credibility pricing blends are designed-for (open-questions §1)."),
    ("Capacity & Exposure", "Underwriting",
     "Utilisation/RAG capacity roll-up + exposure aggregation by any dimension, PML/TIV, peril x geo heatmap, and HARD/SOFT accumulation control at bind time (server capacityMgmt/exposureMgmt; domain).",
     "Delivered",
     "Integer aggregation; bind-time accumulation block (HARD=409 rollback, SOFT=warn+audit) with admin override and RDS evaluation.",
     "Verisk/RMS, Moody's RMS, Sequel",
     "Partial",
     "Peak-zone accumulation is core to cat-modelling vendors; RI admin systems track limits more coarsely.",
     "Live cat-model vendor adapters and clash analysis are designed-for (gap-analysis Tier 4)."),
    ("Territory Workspace", "Underwriting",
     "Cross-module roll-up combining exposure (TIV/PML by country) with geographic capacity in one view (/api/territories).",
     "Delivered",
     "One reconciled view across the exposure and capacity engines; RLS-scoped.",
     "Verisk/RMS, Moody's RMS, Sequel",
     "Varies",
     "Geographic accumulation views are a cat-tool/analytics feature, not universal in admin systems.",
     "Depends on the same cat-model adapters for live hazard data."),
    ("Retrocession", "Underwriting",
     "Quota-share allocation rules (LOB/currency/period filters, priorities) + allocation run (POST /api/retrocession/allocation/run) booking ceded events with a full source->rule->ceded trace.",
     "Partial",
     "Largest-remainder integer allocation capped at source, idempotent via DB constraint, fully traceable.",
     "SAP FS-RI, Sapiens, Effisoft, DXC",
     "Yes",
     "Outward/retro programmes with cession engines are standard in FS-RI-class systems.",
     "Non-QS cession methods are a follow-on; the retro create flow is the thinnest contract form (gap-analysis §1.4)."),
    ("Adjustments", "Underwriting",
     "Premium tracking (EPI/minimum/deposit/booked) and M&D premium adjustment booking max(minimum, rate x GNPI) - booked as an audited event (server treatyAdjustments; domain premiumAdjustment).",
     "Delivered",
     "Idempotent on re-run, return-premium aware, posts through the same financial-event path.",
     "SAP FS-RI, Sapiens, Effisoft",
     "Yes",
     "Minimum-and-deposit / adjustment premium processing is a standard technical-accounting function.",
     "Auto-invoke from period close is a follow-on."),

    # --- Distribution ---
    ("Parties", "Distribution",
     "Party/role core with identifiers (LEI, tax ID, NAIC/Lloyd's), security ratings + credit limits + collateral register, and deterministic sanctions screening (server parties; migrations 0052-0053).",
     "Delivered",
     "Party-role model, integer-exact credit headroom, security-committee view, normalised sanctions matcher screening on create.",
     "SAP FS-RI, Sapiens, Guidewire, Duck Creek, Oracle, SICS",
     "Yes",
     "A counterparty/party master with roles and financial data underpins every platform.",
     "Contacts/bank details are post-creation; group->entity->branch hierarchy and live OFAC/UN/EU feeds are designed-for (gap-analysis §1.2)."),
    ("Clients", "Distribution",
     "Client list/analytics view over the party master (server clients).",
     "Partial",
     "Projection over the single party-role core, so no duplicate master.",
     "Novidea, Salesforce-based insurance CRM, SAP",
     "Varies",
     "Client/account views are usually a CRM or party-master facet, not a distinct RI module.",
     "Full account 360/CRM depth is partial."),
    ("Brokers", "Distribution",
     "Broker analytics: profitability, transparent relationship score/tiers, derived GWP, hierarchy/group structure (server brokers; domain counterparty).",
     "Delivered",
     "Transparent relationship score and bands; deep-links into the UW workbench.",
     "SAP FS-RI, Sapiens, Sequel, Novidea",
     "Partial",
     "Broker/intermediary management exists in admin suites and distribution platforms; analytics depth varies.",
     "Broker/renewal dashboards are designed-for (phases.md UW)."),
    ("Cedents", "Distribution",
     "Cedent analytics: GWP, performance, hierarchy, contracts, loss/claims history, communications (server cedents; domain counterparty).",
     "Delivered",
     "Same transparent scoring model as brokers; consistent party-role core.",
     "SAP FS-RI, Sapiens, Effisoft, SICS",
     "Yes",
     "Cedent (client company) management is core to reinsurer admin systems.",
     "Renewal/portfolio dashboards are designed-for."),
    ("CRM", "Distribution",
     "Relationship-management surface (server crm) over the party master.",
     "Partial",
     "Built on the shared party core rather than a bolt-on CRM silo.",
     "Novidea, Salesforce, Microsoft Dynamics-based insurance CRM",
     "Varies",
     "Distribution CRM is typically a specialised platform (Novidea) or a Salesforce/Dynamics overlay.",
     "Full pipeline/activity CRM is partial."),

    # --- Operations ---
    ("Claims", "Operations",
     "Notify -> reserve movement -> paid-loss financial event, with catastrophe/occurrence coding, reinstatement processing and cash-call workflow (server claims/claimsAdvanced; domain nonproportional).",
     "Delivered",
     "Reserve movements and paid losses are audited events feeding the same reconciling chain; cat/event aggregation; maker/checker cash calls.",
     "Guidewire ClaimCenter, Sapiens, SAP FS-CM, Effisoft, DXC",
     "Yes",
     "Claims/loss administration with reserves and recoveries is a core module everywhere.",
     "The full reference-data-driven claim state machine is only partially wired (open-questions §3)."),
    ("Bordereaux", "Operations",
     "Stored column-mapping projects arbitrary headers onto canonical fields; each line validated; amounts quantised to integer minor units; line sum reconciled to a declared control total (out-of-balance REJECTED). RFC-4180 CSV parsing delivered.",
     "Delivered",
     "Control-total reconciliation rejects out-of-balance files; integer quantisation; line-numbered parse errors.",
     "Sequel, Eurobase, Effisoft, DXC",
     "Yes",
     "Bordereaux ingestion with per-cedent templates is a standard delegated-authority function.",
     "Excel (.xlsx) parsing and streaming/connector ingestion designed-for (open-questions §1); per-cedent column schema library is the gap."),
    ("Recoveries", "Operations",
     "Recovery position nets a claim through its recovery ledger (received vs expected), inuring application order, and event aggregation to occurrence level (server recoveries; domain claimsRecovery).",
     "Delivered",
     "Integer-exact netting, inuring before the protected layer, /api/claims/:id/net-position.",
     "SAP FS-RI, Sapiens, Guidewire, Effisoft",
     "Yes",
     "Outward recovery / inward salvage-subrogation tracking is standard.",
     "Portfolio/programme rollup UI is designed-for (open-questions §3)."),
    ("Operations Center", "Operations",
     "Task + SLA management (server operations/tasks; domain tasks) with referral tasks auto-created from UW.",
     "Partial",
     "SLA state machine tied into UW referrals; audited.",
     "Guidewire, Duck Creek, Sapiens",
     "Yes",
     "Work queues/activity management ship with the major suites.",
     "Broader ops orchestration is partial."),
    ("Workflow Center", "Operations",
     "Metadata-driven workflow engine (server workflowEngine; domain workflow) executing stage transitions.",
     "Partial",
     "Config-driven stages rather than hard-coded; audited transitions.",
     "Guidewire, Duck Creek, Fadata, Pega-class",
     "Yes",
     "Author-time workflow/process configuration is central to Guidewire/Duck Creek.",
     "The no-code workflow designer UI is designed-for (open-questions §6)."),
    ("Audit Log", "Operations",
     "Read-only viewer over the hash-chained, append-only audit_log (server auditLog, /api/audit) with entity/action facets.",
     "Delivered",
     "Hash-chained tamper-evident chain; the app DB role has NO UPDATE/DELETE on audit_log; every material mutation writes in-transaction.",
     "SAP FS-RI, Sapiens, Guidewire, Oracle",
     "Partial",
     "All keep audit trails, but cryptographic hash-chaining with no-delete enforcement at the DB is a genuine differentiator.",
     "Extending audit to every config write is a small hardening item (open-questions §5)."),

    # --- Finance ---
    ("Accounting", "Finance",
     "Double-entry GL: immutable financial events -> statement -> balanced journal postings; reconcile() proves the control-account movement equals the statement balance (reconciles to zero).",
     "Delivered",
     "Integer minor units end to end; the reconciling technical->financial chain is the foundation; balanced postings; per-type posting rules.",
     "SAP FS-RI, Sapiens, FIS, Oracle Insurance, Effisoft",
     "Yes",
     "Technical + financial accounting with a GL is the heart of every RI admin/ERP system.",
     "Configurable posting-rule reference data is the full-platform extension; a controlled manual-journal screen is a standard add (gap-analysis §1.4)."),
    ("Statements", "Finance",
     "Statement-of-account lifecycle (Open->Prepared->...->Settled) with AR/AP invoice spin-off on issue, and a verifier that recomputes commissions/brokerage/reinstatement from the typed terms and flags deviations (server statements/soaVerification).",
     "Delivered",
     "Terms-recomputation verification (the FS-RI-class differentiator) with tolerance flags; guarded transitions; unverifiable items fail loudly.",
     "SAP FS-RI, Sapiens ReinsuranceMaster, Eurobase, Effisoft",
     "Yes",
     "Technical-account statements verified against contract terms are the reinsurance-system core function.",
     "Profit-commission jurisdictional variants beyond the common basis are configuration (open-questions §2)."),
    ("Finance", "Finance",
     "Account-current per party (open AR/AP, net per currency, aging, dunning), disputed items that pause dunning, and payment runs with maker-checker release generating ISO 20022 pain.001 (server finance/accountCurrent; domain aging).",
     "Delivered",
     "Exact-decimal control sum on the pain.001 file; maker-checker release; cash booking stays on the financial-event path.",
     "SAP FS-RI, FIS, Oracle, Sapiens",
     "Yes",
     "AR/AP, account-current netting and bank-file generation are standard finance functions.",
     "The bank file is produced; live bank connectivity is a deployment integration (gap-analysis Tier 2)."),
    ("Treasury", "Finance",
     "Amortised-cost / effective-interest engine (yield-to-maturity by bisection, IFRS 9 amortisation to par), accrued interest and portfolio valuation (server treasury; domain treasury/amortisedCost).",
     "Partial",
     "IFRS 9 amortised-cost schedule converging to par; integer-exact.",
     "FIS, Oracle, SimCorp-class investment systems",
     "Partial",
     "Investment accounting exists in finance suites; a full dealing/settlement sub-ledger is specialist.",
     "Cash-flow forecasting and a full dealing/settlement sub-ledger + market data are designed-for (open-questions §2)."),
    ("Period Close", "Finance",
     "Period-close surface (server periodClose) coordinating valuation runs incl. the UPR/DAC earning run (POST /api/accounting/upr/run).",
     "Partial",
     "UPR earning is integer-exact (earned + UPR = written) across pro-rata/8ths/24ths/risk-attaching.",
     "SAP FS-RI, FIS, Oracle, Sapiens",
     "Yes",
     "Period/year-end close orchestration is standard ERP/RI functionality.",
     "Auto-invoking the UPR run from the close and full close orchestration are follow-ons (gap-analysis Tier 2)."),
    ("Procurement", "Finance",
     "Procurement surface (server procurement) for purchase/expense capture.",
     "Partial",
     "Shares the tenant, RBAC and audit foundation.",
     "SAP (ERP), Oracle",
     "Varies",
     "Procurement is a general-ERP module, not a reinsurance-specific one; only broad ERPs cover it.",
     "Peripheral to the RI core; kept thin."),

    # --- Analytics & Compliance ---
    ("Reports", "Analytics & Compliance",
     "Governed report-pack assembler (server reporting; domain reportPack): template-driven sections/line items with computed totals, completeness checks and control-total tie-outs.",
     "Partial",
     "Definitions-as-data with control-total tie-outs, resolved regardless of declaration order.",
     "SAP FS-RI, Sapiens, Moody's RMS, FIS",
     "Yes",
     "Reporting sits on the vendor BI layer or an external DW/semantic layer.",
     "Drag-drop report designer, Excel/PDF packs and a semantic layer are designed-for (open-questions §9)."),
    ("Scheduled Reports", "Analytics & Compliance",
     "Report cadence scheduling (server scheduledReports/scheduler; domain reportCadence).",
     "Partial",
     "Deterministic cadence computation, reusing the shared scheduler.",
     "SAP FS-RI, Sapiens, Oracle",
     "Yes",
     "Report scheduling/distribution is standard BI functionality.",
     "Distribution to external sinks is provider-wired."),
    ("Analytics", "Analytics & Compliance",
     "Loss analytics incl. chain-ladder IBNR, loss ratio, frequency/severity, technical account (server analytics; domain analytics/lossAnalytics).",
     "Partial",
     "Actuarial engines (chain-ladder) are pure and unit-tested.",
     "Verisk, Moody's RMS, Guidewire, FIS",
     "Partial",
     "Deep analytics is an analytics-vendor or BI-layer capability; RI admin systems vary.",
     "Pivot/cube/drill-through BI is designed-for (open-questions §9)."),
    ("Risk & Capital", "Analytics & Compliance",
     "Solvency II standard-formula Pillar 1 (premium/reserve risk, SCR aggregation via correlation matrix, BSCR+OpRisk+LAC, MCR corridor, risk margin, own-funds tiering) and ORSA projection/stress (server riskCapital; domain solvency2/orsa).",
     "Partial",
     "Correlation-matrix SCR, cost-of-capital risk margin, own-funds tiering with solvency ratio, ORSA roll-forward with breach flags - all unit-tested.",
     "Moody's RMS, FIS, msg global, Oracle",
     "Partial",
     "SII/capital modelling is a specialist regulatory/actuarial capability (msg global, Moody's) or an internal-model engine.",
     "Full sub-module granularity, official correlation matrices and certified QRTs are designed-for (open-questions §4)."),
    ("Regulatory", "Analytics & Compliance",
     "Jurisdiction packs assembled from live data (server regulatory/regulatoryAdvanced/jurisdictionPacks): NAIC Schedule F (configurable provision calc), Solvency II QRT skeletons (S.02.01/S.31.01), IRDAI returns - each labelled template, not certified content.",
     "Partial",
     "Packs bind to live tenant data with control-total tie-outs and are honestly labelled template-not-filing.",
     "SAP FS-RI, msg global, Sapiens, specialist reg-reporting tools",
     "Partial",
     "Certified regulatory reporting with official taxonomies/factor tables and filing validation is a specialist capability.",
     "Certified line taxonomies, official factor tables and filing validations remain jurisdiction-specific configuration (open-questions §4)."),
    ("Compliance", "Analytics & Compliance",
     "Compliance surface (server compliance) incl. sanctions screening (denylist matcher, screening log, screen-on-create).",
     "Partial",
     "Deterministic normalised matcher (BLOCKED/POTENTIAL/CLEAR) with an auditable screening log.",
     "SAP FS-RI, Guidewire, Oracle, specialist KYC/AML tools",
     "Yes",
     "Compliance/AML checks are standard, usually via a specialist screening provider.",
     "Live OFAC/UN/EU provider feeds populate the list per deployment (gap-analysis Tier 1)."),
    ("Returns", "Analytics & Compliance",
     "Statutory/market returns via the financial-statements + report-pack engines (server returns/financialStatements).",
     "Partial",
     "Assembled from live data with tie-outs; labelled template.",
     "SAP FS-RI, msg global, FIS",
     "Partial",
     "Statutory return generation is a regulated-reporting capability of finance/reg suites.",
     "Certified returns are configuration; these are templates (open-questions §4)."),

    # --- HRMS ---
    ("Attendance", "HRMS",
     "Enumerated auditable day status (present/absent/leave/holiday/regularized/OD/WFH), monthly grid, and OD/WFH/regularization requests routed to the manager resolved from the org hierarchy (server attendance/hrAttendance; domain attendanceStatus).",
     "Delivered",
     "Manager-as-approver resolved via the recursive org hierarchy (not flat hr:write); regularization keeps original punches; hash-chained status history.",
     "Workday, SAP SuccessFactors, Oracle HCM (HR suites)",
     "No",
     "Attendance is an HRMS-suite capability; dedicated reinsurance platforms do not include it.",
     "Time-based auto-escalation to skip-level manager is designed-for (open-questions §15)."),
    ("People", "HRMS",
     "Employee master with employment type, audited status lifecycle (active/on-leave/suspended/exited) and an org-chart rollup via recursive CTE (server hrms).",
     "Delivered",
     "Hash-chained employee_status_history; system roles from the Permission Engine surfaced alongside HR designation.",
     "Workday, SAP SuccessFactors, Oracle HCM",
     "No",
     "Core HR / employee master is an HRMS capability, not an RI-platform one.",
     "Self-service login provisioning for HR-created employees is designed-for (open-questions §15)."),
    ("Payroll", "HRMS",
     "Payroll computation surface (server payroll; domain payroll).",
     "Partial",
     "Pure payroll math on the shared money primitives.",
     "SAP Payroll, Oracle HCM, ADP",
     "No",
     "Payroll is a dedicated HR/payroll-suite function; not present in RI admin systems.",
     "Peripheral to the RI core."),
    ("Performance", "HRMS",
     "Performance-management surface (server performance; domain performance).",
     "Partial",
     "Shares tenant/RBAC/audit foundation.",
     "Workday, SAP SuccessFactors",
     "No",
     "Performance management is an HR-suite function.",
     "Peripheral to the RI core."),
    ("Assets", "HRMS",
     "Fixed-asset register with straight-line and reducing-balance depreciation, net book value and disposal gain/loss (server assets; domain fixedAssets).",
     "Delivered",
     "Depreciation never below residual; integer-exact; disposal gain/loss.",
     "SAP (ERP), Oracle, FIS",
     "Varies",
     "Fixed assets is a general-ledger/ERP module; only broad ERPs include it.",
     "Asset lifecycle beyond depreciation is thin."),
    ("Org Structure", "HRMS",
     "Org units and reporting hierarchy (server organization) backing manager resolution and rollups.",
     "Delivered",
     "Recursive-CTE hierarchy reused across HR approvals and reporting.",
     "Workday, SAP, Oracle HCM",
     "Varies",
     "Org modelling is an HR/ERP capability.",
     "Used chiefly to support HR and delegation."),

    # --- Master Data ---
    ("Products", "Platform",
     "Product lifecycle + metadata product model (server products; domain product). API present (POST /api/products); no authoring screen yet.",
     "Partial",
     "Metadata code lists + product lifecycle + Formula Engine are the same product-definition idea as the vendor suites.",
     "Guidewire, Duck Creek, Fadata, ClarionDoor, Instanda",
     "Yes",
     "Product-definition-driven processing (author-time product model) is the core of Guidewire/Duck Creek/ClarionDoor.",
     "The product studio / authoring UI is the gap (gap-analysis §1.4, §2.3)."),

    # --- Documents & Knowledge ---
    ("Documents", "Platform",
     "Versioned documents with deterministic OCR/AI field extraction (stub), supersede chains and a lightweight e-signature seal (server documents; domain underwritingDocuments/ocr).",
     "Partial",
     "Version + supersede chains and field extraction from text; storage_ref pointer ready for blob storage.",
     "Guidewire, Duck Creek, SAP, Oracle (ECM integrations)",
     "Yes",
     "Document/content management is standard, usually via an ECM integration.",
     "Blob storage and the image/PDF->text OCR step are external/designed-for (open-questions §1, §13)."),

    # --- Integration & Automation ---
    ("Integration Hub", "Platform",
     "Typed connector registry with config validation and secret redaction; transactional message/event outbox + relay (server integration/connectors/eventbus).",
     "Partial",
     "Real, tested orchestration mechanics (outbox, relay, registry, one-time API-key hashing); config-shape validation.",
     "SAP FS-RI, DXC, Sapiens, Guidewire (integration frameworks); ACORD in Sequel/DXC",
     "Partial",
     "Integration frameworks and ACORD messaging exist in mature platforms; depth varies.",
     "ACORD EBOT/ECOT, bureau connectivity and live connector handshakes are designed-for (open-questions §8, §11)."),
    ("Messaging", "Platform",
     "Transactional message outbox with status tracking + notification engine (server messaging/notifications).",
     "Partial",
     "Transactional outbox pattern; dev provider logs/marks sent; production points at a real SMTP/SMS sink.",
     "SAP, Guidewire, Duck Creek",
     "Yes",
     "Notification/correspondence engines are standard.",
     "Email/SMS delivery sink is provider-wired per deployment (open-questions §11)."),
    ("Automation Studio", "AI",
     "Composes the existing rules engine + event bus (trigger -> rule set -> actions), evaluated live (server automation/automationStudio).",
     "Partial",
     "Not a second engine - reuses the tested rules + event-bus primitives.",
     "Guidewire, Duck Creek, Pega-class",
     "Partial",
     "Author-time automation/orchestration is a strength of Guidewire/Duck Creek/Pega.",
     "The visual authoring UI is designed-for (open-questions §6, §14)."),
    ("Portal", "Distribution",
     "Thin scoped projections (server portals): broker/cedent portals see only their own party's contracts/statements/claims via RLS + party scope.",
     "Partial",
     "Portal isolation is enforced by the same DB RLS as the core, scoped to the party - not app-layer filtering.",
     "Sequel, DXC, Instanda, Guidewire (digital/portal)",
     "Partial",
     "Broker/cedent/coverholder portals exist in market platforms; Instanda is portal/product-centric.",
     "Full portal breadth (retro/client/coverholder self-service) is designed-for (open-questions §8)."),

    # --- Administration ---
    ("Admin", "Platform",
     "Tenant, role and configuration administration (server platform/admin) incl. metadata code-list management (add values with no deployment).",
     "Partial",
     "Metadata-driven config: statuses/LOBs/roles are code lists, not hard-coded enums; admin:manage overrides every check.",
     "SAP FS-RI, Sapiens, Guidewire, Duck Creek",
     "Yes",
     "Admin consoles for tenants/roles/config are standard.",
     "No-code config designers/sandbox/promotion are designed-for (open-questions §6)."),
    ("Formula Engine", "Platform",
     "Safe, injection-free expression evaluator; formulas as versioned, effective-dated data with named terms so every value carries a step-by-step breakdown; SYSTEM/OVERRIDE/IMPORTED/MANUAL status; audited override+restore; deterministic explain (server formulas; domain formula/formulaLibrary; migration 0050).",
     "Delivered",
     "The first fully-delivered no-code surface: versioned formulas-as-data, INPUT/CALCULATED/PROTECTED governance, audited override trail, grounded explain.",
     "Duck Creek, Guidewire, ClarionDoor (rating/calc engines)",
     "Partial",
     "Author-time rating/calculation engines are central to Duck Creek/Guidewire/ClarionDoor product configuration.",
     "Wiring CalculatedValue into every screen, an approval workflow and drag-drop authoring are designed-for (open-questions §6)."),
    ("Legal Entities", "Platform",
     "Legal-entity structure + multi-GAAP parallel ledgers (gl_ledger + basis-adjustment layer), per-account trial balance and an intercompany-elimination consolidation VIEW (server organization; multiGaap; migration 0058).",
     "Partial",
     "Core+adjustment parallel-ledger model; the single-ledger GL is provably untouched.",
     "SAP (ERP), Oracle, FIS",
     "Yes",
     "Multi-entity / multi-GAAP is a finance-suite capability.",
     "Honestly a consolidation VIEW, not a legal-entity consolidation engine (gap-analysis Tier 3)."),
    ("Ops Console", "Platform",
     "Operational monitoring surface (server operations).",
     "Partial",
     "Shares tenant/RBAC/audit foundation.",
     "Guidewire, Duck Creek (ops tooling)",
     "Varies",
     "Ops consoles vary by platform and are often external observability tools.",
     "Full metrics/traces/SLO observability is designed-for (open-questions §7)."),
    ("Delegation", "Platform",
     "Authority delegation (server delegation; domain delegation) - delegate permissions/approvals for a period.",
     "Partial",
     "Reuses RBAC; audited.",
     "Sequel (binding authority), SAP, Guidewire",
     "Partial",
     "Delegated/binding authority is a coverholder-market concept (Sequel); generic delegation varies.",
     "Binding-authority depth is a follow-on."),
    ("Security", "Platform",
     "JWT/RBAC + DB-enforced RLS, TOTP MFA, OIDC SSO, SAML SP metadata, WebAuthn ceremonies, and KMS envelope encryption (AES-256-GCM) (server security/saml/webauthn/kms).",
     "Partial",
     "DB-enforced tenant isolation (fail-closed), hash-chained audit, real MFA/OIDC, tested envelope encryption.",
     "SAP FS-RI, Guidewire, Duck Creek, Oracle (enterprise security)",
     "Partial",
     "Enterprise auth (SSO/MFA/RBAC) is standard; RLS at the database is a stronger-than-typical isolation model.",
     "SAML signature validation, WebAuthn attestation and a managed HSM/KMS master key are provider-wired (open-questions §5, §12)."),
    ("Security Ops", "Platform",
     "Security operations surface (server securityOps).",
     "Partial",
     "Login rate limiting delivered; sits on the audited foundation.",
     "External SIEM/SOC tooling",
     "Varies",
     "SOC/SIEM is typically a dedicated external platform, not part of the admin suite.",
     "SOC/SIEM integration and SAST/DAST/pen-test are designed-for (open-questions §5)."),
    ("Field Security", "Platform",
     "Field-level security / column masking surface (server fieldSecurity; domain masking).",
     "Designed-for",
     "Masking primitives exist in the domain; enforcement in queries is not yet applied.",
     "Salesforce, Guidewire, Oracle",
     "Partial",
     "Field-level security exists in configurable platforms; not universal.",
     "FLS / column masking enforcement is designed-for; ABAC modelled not applied (open-questions §5)."),
    ("Retention", "Platform",
     "Data-retention surface (server retention).",
     "Designed-for",
     "Reuses tenant/audit foundation.",
     "SAP, Oracle, records-management tools",
     "Partial",
     "Records retention/legal hold is a governance capability of mature suites.",
     "Retention / legal hold / right-to-erasure are designed-for (open-questions §5)."),
    ("Cost Management", "Platform",
     "Cost/usage surface (server cost).",
     "Partial",
     "Tenant-scoped usage figures.",
     "Cloud FinOps / metering tools",
     "Varies",
     "Usage/cost metering is a platform/FinOps concern, not an RI-admin one.",
     "Peripheral; usage metering depth is thin."),
    ("Features", "Platform",
     "Feature-flag surface (server platform features).",
     "Designed-for",
     "Reuses tenant config.",
     "SaaS entitlement/flag platforms (LaunchDarkly-class)",
     "Varies",
     "Entitlement/feature flagging is a SaaS-platform capability.",
     "The per-tenant/plan entitlement engine (flags & limits) is designed-for (open-questions §6)."),
]

# ---------------------------------------------------------------------------
# SHEET 2 — Engines
# ---------------------------------------------------------------------------
ENGINE_HEADERS = [
    "Engine", "What it does in RIOS", "Status", "Where it lives (code)",
    "Comparator equivalent", "Notes / gaps",
]

ENGINES = [
    ("Dynamic Form Renderer",
     "Metadata-driven form rendering: a slip/form renders and validates against a typed model definition so a new model is a data change, not a redeploy. config_document store exists.",
     "Designed-for",
     "web form renderer (in progress); config_document store; domain underwritingModels",
     "Guidewire/Duck Creek author-time form & UI configuration; Instanda no-code product forms",
     "The metadata-driven form renderer (brief S10.3) is a key designed-for item; interpreters/designer UIs designed-for (phases.md Ph.9, open-questions S6)."),
    ("Workflow Engine",
     "Executes config-defined stage transitions and lifecycle state machines with illegal-transition rejection; auto-creates referral tasks.",
     "Partial",
     "server workflowEngine; domain workflow; lifecycle guards across treaty/UW/claims",
     "Guidewire/Duck Creek workflow & process authoring; Pega-class BPM",
     "Engine delivered and tested; the no-code workflow designer UI is designed-for (open-questions S6)."),
    ("Formula Engine",
     "Safe injection-free expression evaluator; formulas as versioned, effective-dated data with named terms and step-by-step breakdowns; SYSTEM/OVERRIDE/IMPORTED/MANUAL + INPUT/CALCULATED/PROTECTED governance; audited override+restore; grounded explain.",
     "Delivered",
     "domain formula/formulaLibrary; server formulas; migration 0050; web Formula Management + CalculatedValue",
     "Duck Creek / Guidewire rating & calculation engines; ClarionDoor rating",
     "First fully-delivered no-code surface. Wiring into every screen, approval workflow and drag-drop authoring are designed-for."),
    ("Validation Engine",
     "Column-to-field mapping with typed coercion and per-cell error reporting (mapAndValidate); typed contract-terms schema with cross-field refinements; bordereaux control-total reconciliation.",
     "Delivered",
     "domain dataImport/rules; server import/validate; typed treaty terms schema",
     "Product-model validation in Guidewire/Duck Creek; bordereaux validation in Sequel/Eurobase",
     "Per-cedent column-schema library is the content gap (gap-analysis S1.4)."),
    ("Document Engine",
     "Versioned documents, supersede chains, deterministic field extraction from text, and an e-signature seal; storage_ref pointer for blob storage.",
     "Partial",
     "server documents; domain underwritingDocuments/ocr",
     "ECM integrations behind Guidewire/Duck Creek/SAP/Oracle",
     "Blob storage and the image/PDF->text OCR step are external/designed-for (open-questions S1, S13)."),
    ("Notification Engine",
     "Transactional message outbox with status tracking; notification composition; event outbox + relay pattern.",
     "Partial",
     "server notifications/messaging/eventbus",
     "Correspondence/notification engines in SAP/Guidewire/Duck Creek",
     "Email/SMS and event-bus sinks are provider-wired per deployment (open-questions S11)."),
    ("Dashboard Engine",
     "Tenant-scoped KPI summary (/api/dashboard/summary) over the reconciling financial data.",
     "Partial",
     "server dashboard/executive/analytics",
     "Vendor BI/analytics layers; external DW + semantic layer",
     "Drag-drop dashboard designer, pivot/cube and drill-through are designed-for (open-questions S9)."),
    ("AI Recommendation Engine",
     "Deterministic, grounded UW advisor (clause recommendations, missing-info detection, consistency flags, exec summary, similar-risk benchmarking) plus the confirm-before-mutate assistant; no LLM required.",
     "Delivered",
     "domain underwritingAdvisor/insight/prediction; server assistant/aiInsights",
     "Limited native AI in RI admin systems; analytics vendors (Verisk/Moody's) supply ML separately",
     "Deterministic core delivered and tested; optional LLM narration/generation is designed-for (open-questions S13, ADR 0005)."),
]

# ---------------------------------------------------------------------------
# SHEET 3 — Scorecard
# ---------------------------------------------------------------------------
SCORE_HEADERS = ["Dimension", "RIOS status", "One-line justification (grounded)",
                 "Comparator footprint"]

SCORECARD = [
    ("Multi-tenancy", "Delivered",
     "DB-enforced RLS via runAs() sets app.tenant_id/app.user_id LOCAL; fail-closed; a negative cross-tenant isolation test proves reads are blocked (open-questions S5, ADR 0002).",
     "Most suites use app-layer or schema tenancy; DB RLS is stronger than typical."),
    ("Money integrity", "Delivered",
     "Integer minor units end to end; same-currency-only arithmetic; cross-currency throws (goes through FX); *_minor bigint columns (ADR 0003).",
     "Established RI/finance systems handle money robustly; integer-only-by-construction is a strong discipline."),
    ("Audit", "Delivered",
     "Hash-chained, append-only audit_log written in-transaction on mutations; the app DB role has NO UPDATE/DELETE on it.",
     "All keep audit trails; cryptographic hash-chaining with DB-enforced immutability is a differentiator."),
    ("Reinsurance math", "Delivered",
     "Proportional/non-proportional, sliding scale (stepped + interpolated), profit commission, reinstatements, swing rating, commutation/LPT, indexation/hours clauses - pure and unit-tested in @rios/domain (461 domain tests).",
     "Matches the FS-RI/Sapiens/Effisoft technical-math footprint."),
    ("Accounting / GL", "Delivered",
     "Double-entry: financial event -> statement -> balanced postings -> reconcile to zero; multi-GAAP core+adjustment parallel ledgers; UPR earning (earned+UPR=written).",
     "Core of SAP FS-RI / Sapiens / FIS; RIOS matches the reconciling-chain model."),
    ("IFRS 17 / Solvency II", "Partial",
     "Measurement engines delivered and unit-tested (PAA/GMM/VFA/CSM roll-forward; SII Pillar 1 SCR/MCR/risk-margin/own-funds; ORSA projection/stress). Persistence/disclosure, full sub-module granularity and certified QRTs are designed-for (open-questions S4).",
     "Specialist capability of msg global / Moody's / FIS; RIOS has the math, not the certified reporting."),
    ("Regulatory packs", "Partial",
     "NAIC Schedule F (configurable provision), SII QRT skeletons (S.02.01/S.31.01), IRDAI returns - assembled from live data, tie-out checked, labelled template-not-certified. Certified taxonomies/factors are configuration (open-questions S4).",
     "Certified regulatory reporting is a specialist reg-tool/finance-suite function."),
    ("Documents", "Partial",
     "Versioned documents, supersede chains, field extraction from text and an e-sign seal delivered; blob storage and image/PDF->text OCR are external/designed-for (open-questions S1, S13).",
     "ECM/DMS integration is standard in mature suites."),
    ("AI", "Partial",
     "Guardrailed assistant: deterministic intent engine, grounded in tenant data, confirm-before-mutate re-checking permissions, no backdoor, fully usable with AI disabled. LLM narration/generation designed-for (ADR 0005).",
     "Native AI is uncommon in RI admin systems; RIOS's guardrailed, disable-able design is distinctive."),
    ("Adaptive forms", "Partial",
     "Formula Engine is the first fully-delivered no-code surface (versioned formulas-as-data, governance, audited overrides). The metadata-driven form renderer is in progress/designed-for (phases.md Ph.9).",
     "Author-time configuration is the Guidewire/Duck Creek/Instanda strength; RIOS has the engine idea, UIs pending."),
    ("Cat modelling", "Partial",
     "Accumulation/PML/AAL/EP-curve/TVaR via a CatModelProvider adapter (MockCatModel) with bind-time accumulation control; live RMS/Verisk adapters and event-loss-table import are designed-for (open-questions S1, gap-analysis Tier 4).",
     "Cat modelling is owned by Verisk/RMS and Moody's RMS; RIOS integrates behind an interface."),
    ("ACORD / bureau", "Designed-for",
     "The ACORD EBOT/ECOT + placing-message connector framework and bureau connectivity are designed-for; the typed connector registry and outbox mechanics exist (open-questions S8, gap-analysis Tier 4).",
     "ACORD messaging is embedded in London-market platforms (Sequel, DXC)."),
    ("Investment sub-ledger", "Partial",
     "Amortised-cost/EIR, accrued interest and portfolio valuation engines delivered; a full dealing/settlement sub-ledger, market data and ALM are designed-for (open-questions S2).",
     "Full investment sub-ledgers live in FIS/Oracle/SimCorp-class systems."),
]

# ---------------------------------------------------------------------------
# SHEET 4 — Legend & Honesty Note
# ---------------------------------------------------------------------------
LEGEND_ROWS = [
    ("Delivered", "Built, working, and tested (unit and/or server-integration). Grounded in server/src/modules/*, packages/domain/src/* and the test suite."),
    ("Partial", "A working core is delivered and tested, but named extensions (breadth, UI, external wiring, or content) remain. The specific gap is stated per row."),
    ("Designed-for", "The architecture/schema permits it and interfaces may exist, but it is not built. Named, never silently faked (brief S3.3)."),
    ("Not-yet", "Used only on the Scorecard for a dimension with no delivered core yet."),
]

HONESTY_PARAGRAPHS = [
    "How the RIOS column was grounded. Every RIOS claim in this workbook is taken from the repository, not "
    "from memory or marketing: the module list is enumerated from web/src/app/nav.ts; mechanics are read from "
    "server/src/modules/* (endpoints, tables, posting rules) and packages/domain/src/* (the maths that "
    "actually exists and is unit-tested); and each Status is cross-checked against the honest registers "
    "docs/phases.md, docs/open-questions.md and docs/industry-gap-analysis.md.",

    "How the comparator columns were grounded. Named reinsurance/insurance platforms (SAP FS-RI, Sapiens "
    "ReinsuranceMaster, Eurobase Synergy2, Guidewire, Duck Creek, Fadata INSIS, Oracle Insurance, FIS, "
    "Majesco, SICS, Xuber/DXC, Verisk/RMS, Moody's RMS, Sequel, Effisoft WebXL/Omega, Tia/TietoEVRY, "
    "msg global, Novidea, Instanda, ClarionDoor) do not publish their internal field lists or source. The "
    "comparator descriptions here are inferred ONLY from public, market-standard footprints: vendor module "
    "maps, the Lloyd's/LMA Market Reform Contract (MRC), ACORD messaging (EBOT/ECOT/placing), NAIC Schedule F, "
    "Solvency II QRTs, and published product-model descriptions. No proprietary competitor internals are "
    "asserted or invented.",

    "What RIOS is. RIOS is a correct, secure, audited vertical-slice FOUNDATION - place -> bind -> account -> "
    "reconcile, plus claims and a guardrailed assistant - not a finished commercial product. The breadth "
    "targeted by the brief (40+ modules, IFRS 17 / Solvency II certified reporting, microservices, full "
    "portals, no-code designers) is designed-for and named, never pretended. This workbook is deliberately "
    "conservative: where a module has a working engine but pending UI, external wiring, or certified content, "
    "it is marked Partial or Designed-for with the specific gap stated.",

    "Deliberate honesty calls (marked down for accuracy): Facultative, Retrocession, Period Close, Treasury, "
    "Documents, Integration Hub, Automation Studio, Portal, Products and Reports/Analytics are Partial "
    "(engine or core present, breadth/UI/external wiring pending); Risk & Capital and Regulatory are Partial "
    "and every regulatory pack is labelled TEMPLATE, NOT CERTIFIED CONTENT; IFRS 17 / Solvency II are Partial "
    "(measurement math delivered, persistence/disclosure/certified QRTs designed-for); Cat modelling and "
    "Investment sub-ledger are Partial (engines present, vendor adapters/dealing sub-ledger designed-for); "
    "ACORD/bureau, Field Security, Retention and Features are Designed-for. HRMS modules (Attendance, People, "
    "Payroll, Performance) are honestly noted as outside the reinsurance core - their comparators are HR "
    "suites (Workday/SuccessFactors), not RI platforms.",
]


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
def build():
    wb = Workbook()

    # Sheet 1 — Module Comparison
    ws1 = wb.active
    ws1.title = "Module Comparison"
    ws1.append(MODULE_HEADERS)
    for row in MODULES:
        ws1.append(list(row))
    style_header(ws1)
    finalize(ws1, widths=[20, 20, 46, 13, 42, 30, 16, 42, 42], status_col=4)

    # Sheet 2 — Engines
    ws2 = wb.create_sheet("Engines")
    ws2.append(ENGINE_HEADERS)
    for row in ENGINES:
        ws2.append(list(row))
    style_header(ws2)
    finalize(ws2, widths=[24, 52, 14, 40, 42, 46], status_col=3)

    # Sheet 3 — Scorecard
    ws3 = wb.create_sheet("Scorecard")
    ws3.append(SCORE_HEADERS)
    for row in SCORECARD:
        ws3.append(list(row))
    style_header(ws3)
    finalize(ws3, widths=[26, 15, 66, 52], status_col=2)

    # Sheet 4 — Legend & Honesty Note
    ws4 = wb.create_sheet("Legend & Honesty Note")
    ws4["A1"] = "RIOS Module Comparison - Legend & Honesty Note"
    ws4["A1"].font = TITLE_FONT
    r = 3
    ws4.cell(r, 1, "Status legend").font = Font(bold=True, size=12)
    r += 1
    ws4.cell(r, 1, "Status").font = HEADER_FONT
    ws4.cell(r, 2, "Meaning").font = HEADER_FONT
    for c in (1, 2):
        ws4.cell(r, c).fill = HEADER_FILL
        ws4.cell(r, c).alignment = CENTER
    r += 1
    for status, meaning in LEGEND_ROWS:
        ws4.cell(r, 1, status)
        fill = STATUS_FILL.get(status)
        if fill:
            ws4.cell(r, 1).fill = fill
        ws4.cell(r, 1).alignment = CENTER
        m = ws4.cell(r, 2, meaning)
        m.alignment = WRAP_TOP
        r += 1

    r += 1
    ws4.cell(r, 1, "Honesty basis").font = Font(bold=True, size=12)
    r += 1
    for para in HONESTY_PARAGRAPHS:
        ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cell = ws4.cell(r, 1, para)
        cell.alignment = WRAP_TOP
        cell.font = NOTE_FONT
        ws4.row_dimensions[r].height = 92
        r += 1

    ws4.column_dimensions["A"].width = 24
    ws4.column_dimensions["B"].width = 118

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "RIOS-module-comparison.xlsx")
    wb.save(out)
    print(f"Wrote {out}")
    print(f"  Module Comparison rows: {len(MODULES)}")
    print(f"  Engines rows:           {len(ENGINES)}")
    print(f"  Scorecard rows:         {len(SCORECARD)}")
    return out


if __name__ == "__main__":
    build()
